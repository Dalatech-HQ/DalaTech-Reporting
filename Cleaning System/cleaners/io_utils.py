from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


MASTER_COLUMNS = [
    "Brand Partners",
    "Particulars",
    "Date",
    "Retailers",
    "Vch Type",
    "Vch No.",
    "Quantity",
    "Sales_Value",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
ALT_FILL = PatternFill(fill_type="solid", fgColor="EBF3FB")
WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=9, color="000000")
THIN_SIDE = Side(style="thin", color="D9E2F3")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

DEFAULT_WIDTHS = {
    "Brand Partners": 20,
    "Particulars": 42,
    "Date": 12,
    "Retailers": 45,
    "Vch Type": 22,
    "Vch No.": 12,
    "Quantity": 10,
    "Sales_Value": 14,
}


def ensure_directory(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def coerce_master_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=MASTER_COLUMNS)

    out = df.copy()
    rename_map = {}
    lower_lookup = {str(c).strip().lower(): c for c in out.columns}
    for required in MASTER_COLUMNS:
        source = lower_lookup.get(required.lower())
        if source is not None and source != required:
            rename_map[source] = required
    if rename_map:
        out = out.rename(columns=rename_map)

    for column in MASTER_COLUMNS:
        if column not in out.columns:
            out[column] = pd.NA

    out = out[MASTER_COLUMNS].copy()
    return out


def write_styled_excel(df: pd.DataFrame, path: Path, sheet_name: str = "Sheet1") -> Path:
    path = Path(path)
    ensure_directory(path.parent)

    frame = coerce_master_columns(df)
    frame.to_excel(path, index=False, sheet_name=sheet_name)

    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, column_name in enumerate(MASTER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = DEFAULT_WIDTHS.get(column_name, 14)

    for row_idx in range(2, ws.max_row + 1):
        row_fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, column_name in enumerate(MASTER_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if column_name in {"Date", "Vch Type", "Vch No.", "Quantity", "Sales_Value"}:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
            if column_name == "Date" and cell.value is not None:
                cell.number_format = "DD-MMM-YY"
            elif column_name == "Quantity" and cell.value is not None:
                cell.number_format = "#,##0.##"
            elif column_name == "Sales_Value" and cell.value is not None:
                cell.number_format = "#,##0.00"

    wb.save(path)
    return path


def save_outputs(output_map: dict[str, pd.DataFrame], output_dir: Path) -> dict[str, Path]:
    paths: dict[str, Path] = {}
    for filename, df in output_map.items():
        paths[filename] = write_styled_excel(df, output_dir / filename)
    return paths


def concat_frames(frames: Iterable[pd.DataFrame]) -> pd.DataFrame:
    cleaned = [coerce_master_columns(frame) for frame in frames if frame is not None and not frame.empty]
    if not cleaned:
        return pd.DataFrame(columns=MASTER_COLUMNS)
    return pd.concat(cleaned, ignore_index=True)

