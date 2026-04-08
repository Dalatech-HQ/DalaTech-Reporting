import json
import os
from datetime import datetime

import pandas as pd
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .brand_names import brand_match_terms
from .period_comparison import build_period_comparison, comparison_basis_label, compare_same_type_reports
from .pdf_generator_html import render_pdf_bytes


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_DIR = os.path.join(BASE_DIR, 'templates')

jinja_env = Environment(
    loader=FileSystemLoader(TEMPLATE_DIR),
    autoescape=select_autoescape(['html', 'xml'])
)


def _safe_float(value, default=0.0):
    try:
        if value is None or value == '':
            return default
        return float(value)
    except Exception:
        return default


def _safe_int(value, default=0):
    try:
        if value is None or value == '':
            return default
        return int(value)
    except Exception:
        return default


def _iso_to_human(value):
    if not value:
        return 'Unknown'
    try:
        return datetime.strptime(str(value), '%Y-%m-%d').strftime('%d %b %Y')
    except Exception:
        return str(value)


def _compact_date_label(start_date, end_date):
    if not start_date or not end_date:
        return 'Date range unavailable'
    return f"{_iso_to_human(start_date)} - {_iso_to_human(end_date)}"


def _sanitize_filename(value):
    text = ''.join(c if c.isalnum() or c in '-_ ' else '_' for c in (value or '').strip())
    return '_'.join(text.split()) or 'Activity_Report'


def _humanize_metric_label(value):
    text = str(value or '').replace('_', ' ').strip()
    if not text:
        return 'Unspecified'
    return text.title()


def _humanize_issue_type(value):
    mapping = {
        'expiry_issue': 'Expiry issue',
        'packaging_issue': 'Packaging issue',
        'credit_note': 'Credit note',
        'competitor_issue': 'Competitor issue',
        'general_feedback': 'General feedback',
        'rejection_issue': 'Rejection issue',
        'out_of_stock': 'Out of stock',
        'opportunity': 'Opportunity',
    }
    return mapping.get(str(value or '').strip().lower(), _humanize_metric_label(value))


LEGACY_ACTIVITY_EXPORT_COLUMNS = [
    'Activity Date',
    'Retailer Name',
    'Retailer City',
    'Close to Expiry Concerns?',
    'Competitor Feedback Notes?',
    'Notes Regarding Expiry?',
    'Notes Regarding Packaging Quality?',
    'Opportunities or Concerns?',
    'Order Generated?',
    'Picture of Shelf',
    'Product feedback for which item?',
    'Shelf Facings?',
    'Shelf Level?',
    'Store Inventory?',
]


def _trend_direction(change):
    if change is None:
        return 'neutral'
    if change > 0:
        return 'up'
    if change < 0:
        return 'down'
    return 'neutral'


def _format_delta(change):
    if change is None:
        return 'n/a'
    sign = '+' if change > 0 else ''
    return f"{sign}{change:.1f}%"


def _build_comparison_card(title, basis, metrics, label=None):
    return {
        'title': title,
        'basis': basis,
        'label': label or 'Not available',
        'metrics': [
            {'label': 'Activities', 'value': _format_delta(metrics.get('activities_change'))},
            {'label': 'Issues', 'value': _format_delta(metrics.get('issues_change'))},
            {'label': 'Opportunities', 'value': _format_delta(metrics.get('opportunities_change'))},
            {'label': 'Stores', 'value': _format_delta(metrics.get('stores_change'))},
        ],
    }


def _parse_summary_json(batch_row):
    raw = (batch_row or {}).get('summary_json') or '{}'
    try:
        return json.loads(raw)
    except Exception:
        return {}


def _get_activity_batch_context(ds, report_id=None, batch_id=None):
    batches = ds.get_activity_batches(limit=500)
    batch = None
    if batch_id:
        batch = next((row for row in batches if int(row.get('id') or 0) == int(batch_id)), None)
    elif report_id:
        batch = next((row for row in batches if row.get('report_id') == report_id), None)
    if not batch:
        return None, {}
    summary = _parse_summary_json(batch)
    return batch, summary


def _source_labels(batch_row, summary):
    source_type = (batch_row or {}).get('source_type') or (summary.get('source_meta') or {}).get('source_type')
    quality_flags = list(summary.get('quality_flags') or [])
    if source_type == 'zip_bundle':
        source_type_label = 'Cleaned weekly zip'
    elif source_type == 'text':
        source_type_label = 'Raw activity export'
    else:
        source_type_label = 'Imported activity source'

    if 'salesperson_metadata_missing_in_cleaned_zip' in quality_flags:
        quality_label = 'Cleaned pack - salesperson detail unavailable'
    elif source_type == 'zip_bundle':
        quality_label = 'Cleaned pack'
    else:
        quality_label = 'Full fidelity'
    return source_type_label, quality_label, quality_flags


def _build_survey_filter(brand_name):
    terms = []
    seen = set()
    for term in brand_match_terms(brand_name):
        compact = term.replace(' ', '')
        for candidate in (term, compact):
            key = candidate.strip().lower()
            if not key or key in seen:
                continue
            seen.add(key)
            terms.append(candidate)

    clauses = []
    params = []
    for term in terms:
        clauses.append("LOWER(COALESCE({alias}.survey_name, '')) LIKE LOWER(?)")
        params.append(f"%{term}%")

    if not clauses:
        clauses.append("LOWER(COALESCE({alias}.survey_name, '')) LIKE LOWER(?)")
        params.append(f"%{(brand_name or '').strip()}%")
    return clauses, params


def _read_activity_frames(ds, brand_name, report_id=None, batch_id=None):
    clauses, params = _build_survey_filter(brand_name)
    survey_events = ' OR '.join(part.format(alias='ae') for part in clauses)
    survey_visits = ' OR '.join(part.format(alias='av') for part in clauses)
    scope_col = 'batch_id' if batch_id else 'report_id'
    scope_value = batch_id if batch_id else report_id
    if not scope_value:
        raise ValueError('An activity batch or report is required')
    with ds._connect() as conn:
        events_df = pd.read_sql_query(
            f"""
            SELECT ae.activity_date, ae.salesman_name, ae.salesman_code, ae.salesman_designation,
                   ae.reporting_person_name, ae.survey_name, ae.retailer_code, ae.retailer_name,
                   ae.retailer_type, ae.retailer_state, ae.retailer_district, ae.retailer_city,
                   ae.question, ae.label, ae.answer
            FROM activity_events ae
            WHERE ae.{scope_col} = ? AND ({survey_events})
            ORDER BY ae.activity_date DESC, ae.id DESC
            """,
            conn,
            params=[scope_value, *params],
        )
        visits_df = pd.read_sql_query(
            f"""
            SELECT av.activity_date, av.salesman_name, av.survey_name, av.retailer_code, av.retailer_name,
                   av.retailer_state, av.retailer_city, av.event_count, av.issue_count,
                   av.opportunity_count, av.photo_count
            FROM activity_visits av
            WHERE av.{scope_col} = ? AND ({survey_visits})
            ORDER BY av.activity_date DESC, av.id DESC
            """,
            conn,
            params=[scope_value, *params],
        )
        issues_df = pd.read_sql_query(
            f"""
            SELECT activity_date, retailer_code, retailer_name, salesman_name, brand_name, sku_name,
                   issue_type, severity, question, label, answer
            FROM activity_issues
            WHERE {scope_col} = ? AND LOWER(COALESCE(brand_name, '')) = LOWER(?)
            ORDER BY activity_date DESC, id DESC
            """,
            conn,
            params=[scope_value, brand_name],
        )
        mentions_df = pd.read_sql_query(
            f"""
            SELECT activity_date, retailer_code, retailer_name, brand_name, sku_name, source_kind, source_value
            FROM activity_brand_mentions
            WHERE {scope_col} = ? AND LOWER(COALESCE(brand_name, '')) = LOWER(?)
            ORDER BY activity_date DESC, id DESC
            """,
            conn,
            params=[scope_value, brand_name],
        )
    for df in (events_df, visits_df, issues_df, mentions_df):
        if 'activity_date' in df.columns:
            df['activity_date'] = pd.to_datetime(df['activity_date'], errors='coerce')
    return events_df, visits_df, issues_df, mentions_df


def _compute_current_metrics(events_df, visits_df, issues_df, mentions_df):
    non_opportunity_issues = issues_df[issues_df['issue_type'].fillna('').str.lower() != 'opportunity'].copy()
    opportunities_df = issues_df[issues_df['issue_type'].fillna('').str.lower() == 'opportunity'].copy()
    current = {
        'stores_visited': int(visits_df['retailer_name'].nunique()) if not visits_df.empty else 0,
        'activities_logged': int(len(events_df)),
        'distinct_salesmen': int(events_df['salesman_name'].fillna('').replace('', pd.NA).dropna().nunique()) if not events_df.empty else 0,
        'issues_found': int(len(non_opportunity_issues)),
        'opportunities_found': int(max(len(opportunities_df), _safe_int(visits_df['opportunity_count'].sum()) if not visits_df.empty else 0)),
        'images_captured': int(_safe_int(visits_df['photo_count'].sum()) if not visits_df.empty else 0),
        'states_covered': int(events_df['retailer_state'].fillna('').replace('', pd.NA).dropna().nunique()) if not events_df.empty else 0,
        'cities_covered': int(events_df['retailer_city'].fillna('').replace('', pd.NA).dropna().nunique()) if not events_df.empty else 0,
        'survey_rows': int(len(events_df)),
        'brand_mentions': int(len(mentions_df)),
        'visits': int(len(visits_df)),
    }
    return current, non_opportunity_issues, opportunities_df


def _report_lookup(ds):
    rows = ds.get_all_reports()
    enriched = []
    for row in rows:
        item = dict(row)
        try:
            item['_start_dt'] = datetime.strptime(item.get('start_date') or '', '%Y-%m-%d')
        except Exception:
            item['_start_dt'] = None
        enriched.append(item)
    return [r for r in enriched if r['_start_dt'] is not None]


def _find_comparison_reports(ds, report):
    report_type = report.get('report_type') or 'weekly'
    reports = _report_lookup(ds)
    comparison = compare_same_type_reports(reports, report)
    previous = comparison.get('previous')
    trailing = comparison.get('trailing') or []
    same_period_last_year = comparison.get('same_period_last_year')
    return previous, trailing, same_period_last_year


def _period_metrics(ds, brand_name, report_id):
    events_df, visits_df, issues_df, mentions_df = _read_activity_frames(ds, brand_name, report_id=report_id)
    current, non_opportunity_issues, opportunity_issues = _compute_current_metrics(events_df, visits_df, issues_df, mentions_df)
    return {
        'current': current,
        'events_df': events_df,
        'visits_df': visits_df,
        'issues_df': issues_df,
        'non_opportunity_issues': non_opportunity_issues,
        'opportunity_issues': opportunity_issues,
        'mentions_df': mentions_df,
    }


def _calc_pct_change(current, previous):
    if previous in (None, 0):
        return None
    return round(((current - previous) / previous) * 100, 1)


def _average_metric(metric_rows, key):
    if not metric_rows:
        return None
    values = [row['current'][key] for row in metric_rows if row['current'].get(key) is not None]
    if not values:
        return None
    return round(sum(values) / len(values), 1)


def _build_comparisons(ds, brand_name, report):
    previous, trailing, same_period_last_year = _find_comparison_reports(ds, report)
    previous_metrics = _period_metrics(ds, brand_name, previous['id']) if previous else None
    trailing_metrics = [_period_metrics(ds, brand_name, row['id']) for row in trailing] if trailing else []
    yoy_metrics = _period_metrics(ds, brand_name, same_period_last_year['id']) if same_period_last_year else None
    return previous, previous_metrics, trailing, trailing_metrics, same_period_last_year, yoy_metrics


def _build_kpi_cards(current, previous_metrics, trailing_metrics, source_quality):
    trailing_avg = {
        key: _average_metric(trailing_metrics, key)
        for key in ['stores_visited', 'activities_logged', 'issues_found',
                    'opportunities_found', 'images_captured', 'cities_covered']
    }
    previous = previous_metrics['current'] if previous_metrics else {}
    cards = [
        ('Stores Visited', current['stores_visited'], 'stores', 'stores_visited'),
        ('Activities Logged', current['activities_logged'], 'survey rows', 'activities_logged'),
        ('Issues Found', current['issues_found'], 'logged issues', 'issues_found'),
        ('Opportunities Found', current['opportunities_found'], 'logged opportunities', 'opportunities_found'),
        ('Images Captured', current['images_captured'], 'photos', 'images_captured'),
        ('Cities Covered', current['cities_covered'], 'cities', 'cities_covered'),
    ]
    result = []
    for label, value, subtitle, key in cards:
        prev_change = _calc_pct_change(value, previous.get(key))
        trailing_value = trailing_avg.get(key)
        trailing_change = _calc_pct_change(value, trailing_value)
        result.append({
            'label': label,
            'value': value,
            'subtitle': subtitle,
            'prev_change': prev_change,
            'trailing_change': trailing_change,
            'prev_change_label': _format_delta(prev_change),
            'trailing_change_label': _format_delta(trailing_change),
            'trend': _trend_direction(prev_change),
            'note': None,
        })
    return result, trailing_avg


def _build_daily_trend(visits_df):
    if visits_df.empty:
        return {'labels': [], 'activities': [], 'issues': [], 'opportunities': []}
    working = visits_df.copy()
    working['day_label'] = working['activity_date'].dt.strftime('%d %b %Y')
    daily = (
        working.groupby('day_label', dropna=False)
        .agg(
            activities=('activity_date', 'size'),
            issues=('issue_count', 'sum'),
            opportunities=('opportunity_count', 'sum'),
        )
        .reset_index()
    )
    return {
        'labels': daily['day_label'].tolist(),
        'activities': [int(v) for v in daily['activities'].fillna(0)],
        'issues': [int(v) for v in daily['issues'].fillna(0)],
        'opportunities': [int(v) for v in daily['opportunities'].fillna(0)],
    }


def _build_period_rows(visits_df):
    if visits_df.empty:
        return []
    working = visits_df.copy()
    working['date'] = working['activity_date'].dt.strftime('%Y-%m-%d')
    if 'issue_count' not in working.columns:
        working['issue_count'] = 0
    if 'opportunity_count' not in working.columns:
        working['opportunity_count'] = 0
    daily = (
        working.groupby('date', dropna=False)
        .agg(
            activities=('activity_date', 'size'),
            issues=('issue_count', 'sum'),
            opportunities=('opportunity_count', 'sum'),
        )
        .reset_index()
    )
    return daily.to_dict('records')


def _build_mix_sections(non_opportunity_issues, opportunity_issues, visits_df):
    issue_mix_rows = (
        non_opportunity_issues.groupby('issue_type').size().reset_index(name='count')
        .sort_values(['count', 'issue_type'], ascending=[False, True])
        if not non_opportunity_issues.empty else pd.DataFrame(columns=['issue_type', 'count'])
    )
    issue_mix = [
        {'label': _humanize_issue_type(row.issue_type), 'count': int(row['count'])}
        for _, row in issue_mix_rows.head(8).iterrows()
    ]

    opportunity_candidates = []
    if not opportunity_issues.empty:
        working = opportunity_issues.copy()
        label_series = working['label'].fillna('').replace('', pd.NA)
        question_series = working['question'].fillna('').replace('', pd.NA)
        working['opportunity_label'] = label_series.fillna(question_series).fillna('General opportunity')
        opp_mix_rows = (
            working.groupby('opportunity_label').size().reset_index(name='count')
            .sort_values(['count', 'opportunity_label'], ascending=[False, True])
        )
        opportunity_candidates = [
            {'label': str(row['opportunity_label'])[:80], 'count': int(row['count'])}
            for _, row in opp_mix_rows.head(8).iterrows()
        ]
    elif not visits_df.empty and visits_df['opportunity_count'].sum() > 0:
        top_store = (
            visits_df.groupby('retailer_name')['opportunity_count'].sum()
            .sort_values(ascending=False)
            .head(8)
        )
        opportunity_candidates = [
            {'label': str(idx), 'count': int(val)}
            for idx, val in top_store.items() if _safe_int(val) > 0
        ]

    highlights = {
        'most_frequent_issue': issue_mix[0]['label'] if issue_mix else 'No issues logged',
        'most_urgent_issue': issue_mix[0]['label'] if issue_mix else 'No urgent issue flagged',
        'most_repeated_opportunity': opportunity_candidates[0]['label'] if opportunity_candidates else 'No repeated opportunity logged',
    }
    return issue_mix, opportunity_candidates, highlights


def _build_store_breakdown(visits_df, mentions_df):
    if visits_df.empty:
        return {'activity_chart': [], 'issues_chart': [], 'opportunity_chart': [], 'rows': []}
    stores = (
        visits_df.groupby(['retailer_name', 'retailer_city', 'retailer_state'], dropna=False)
        .agg(
            visits=('activity_date', 'size'),
            issues=('issue_count', 'sum'),
            opportunities=('opportunity_count', 'sum'),
            photos=('photo_count', 'sum'),
            last_activity=('activity_date', 'max'),
        )
        .reset_index()
    )
    if not mentions_df.empty:
        mentions = (
            mentions_df.groupby('retailer_name')
            .agg(
                brand_mentions=('brand_name', 'size'),
                sku_mentions=('sku_name', lambda s: int(s.fillna('').replace('', pd.NA).dropna().nunique()))
            )
            .reset_index()
        )
        stores = stores.merge(mentions, on='retailer_name', how='left')
    else:
        stores['brand_mentions'] = 0
        stores['sku_mentions'] = 0
    stores['brand_mentions'] = stores['brand_mentions'].fillna(0).astype(int)
    stores['sku_mentions'] = stores['sku_mentions'].fillna(0).astype(int)
    stores['last_activity'] = stores['last_activity'].dt.strftime('%d %b %Y')
    rows = stores.sort_values(['visits', 'issues', 'opportunities'], ascending=[False, False, False])
    return {
        'activity_chart': [{'label': row['retailer_name'], 'value': int(row['visits'])} for _, row in rows.head(10).iterrows()],
        'issues_chart': [{'label': row['retailer_name'], 'value': int(row['issues'])} for _, row in rows.sort_values(['issues', 'visits'], ascending=[False, False]).head(10).iterrows()],
        'opportunity_chart': [{'label': row['retailer_name'], 'value': int(row['opportunities'])} for _, row in rows.sort_values(['opportunities', 'visits'], ascending=[False, False]).head(10).iterrows()],
        'rows': rows.head(25).to_dict('records'),
    }


def _build_geography(events_df):
    if events_df.empty:
        return {'states': [], 'cities': [], 'top_state': 'No coverage', 'top_city': 'No coverage'}
    state_rows = (
        events_df.groupby('retailer_state', dropna=False)
        .agg(activities=('activity_date', 'size'), stores=('retailer_name', 'nunique'))
        .reset_index()
        .rename(columns={'retailer_state': 'state'})
        .sort_values(['activities', 'state'], ascending=[False, True])
    )
    city_rows = (
        events_df.groupby(['retailer_city', 'retailer_state'], dropna=False)
        .agg(activities=('activity_date', 'size'), stores=('retailer_name', 'nunique'))
        .reset_index()
        .rename(columns={'retailer_city': 'city', 'retailer_state': 'state'})
        .sort_values(['activities', 'city'], ascending=[False, True])
    )
    return {
        'states': state_rows.head(12).to_dict('records'),
        'cities': city_rows.head(12).to_dict('records'),
        'top_state': state_rows.iloc[0]['state'] if not state_rows.empty else 'No coverage',
        'top_city': city_rows.iloc[0]['city'] if not city_rows.empty else 'No coverage',
    }


def _build_field_team(events_df, visits_df, quality_label):
    if visits_df.empty:
        return {'available': False, 'note': 'No field activity rows available.', 'table': [], 'activity_chart': [], 'issue_chart': [], 'store_chart': []}
    salesmen = visits_df['salesman_name'].fillna('').replace('', pd.NA).dropna()
    if salesmen.empty:
        note = 'Field team detail unavailable in cleaned weekly pack' if quality_label != 'Full fidelity' else 'Field team detail unavailable'
        return {'available': False, 'note': note, 'table': [], 'activity_chart': [], 'issue_chart': [], 'store_chart': []}
    rows = (
        visits_df.groupby('salesman_name', dropna=False)
        .agg(
            activities=('activity_date', 'size'),
            stores=('retailer_name', 'nunique'),
            issues=('issue_count', 'sum'),
            opportunities=('opportunity_count', 'sum'),
        )
        .reset_index()
        .sort_values(['activities', 'issues'], ascending=[False, False])
    )
    return {
        'available': True,
        'note': None,
        'table': rows.head(15).to_dict('records'),
        'activity_chart': [{'label': row['salesman_name'], 'value': int(row['activities'])} for _, row in rows.head(10).iterrows()],
        'issue_chart': [{'label': row['salesman_name'], 'value': int(row['issues'])} for _, row in rows.sort_values(['issues', 'activities'], ascending=[False, False]).head(10).iterrows()],
        'store_chart': [{'label': row['salesman_name'], 'value': int(row['stores'])} for _, row in rows.sort_values(['stores', 'activities'], ascending=[False, False]).head(10).iterrows()],
    }


def _build_details(events_df):
    if events_df.empty:
        return {'rows': [], 'count': 0, 'pdf_rows': [], 'flat_rows': [], 'preview_rows': []}
    details = events_df.copy()
    details['activity_date'] = details['activity_date'].dt.strftime('%d %b %Y')
    details = details.fillna('')
    cols = [
        ('activity_date', 'Activity Date'),
        ('salesman_name', 'Salesman Name'),
        ('retailer_name', 'Retailer Name'),
        ('retailer_type', 'Retailer Type'),
        ('retailer_city', 'Retailer City'),
        ('retailer_state', 'Retailer State'),
        ('survey_name', 'Survey Name'),
        ('question', 'Question'),
        ('label', 'Label'),
        ('answer', 'Answer'),
    ]
    ordered = []
    for _, row in details.iterrows():
        ordered.append({label: row[col] for col, label in cols})

    flat_df = events_df.copy()
    flat_df['activity_date'] = flat_df['activity_date'].dt.strftime('%d-%b-%y')
    flat_df = flat_df.fillna('')
    group_cols = ['activity_date', 'retailer_name', 'retailer_city']

    def collapse(values):
        seen = []
        for value in values:
            text = str(value or '').strip()
            if not text or text.lower() == 'none':
                continue
            if text not in seen:
                seen.append(text)
        if not seen:
            return ''
        return seen[0] if len(seen) == 1 else ' | '.join(seen[:3])

    pivoted = (
        flat_df.groupby(group_cols + ['label'], dropna=False)['answer']
        .agg(collapse)
        .reset_index()
        .pivot(index=group_cols, columns='label', values='answer')
        .reset_index()
    )
    pivoted.columns.name = None
    pivoted = pivoted.rename(columns={
        'activity_date': 'Activity Date',
        'retailer_name': 'Retailer Name',
        'retailer_city': 'Retailer City',
    })
    for column in LEGACY_ACTIVITY_EXPORT_COLUMNS:
        if column not in pivoted.columns:
            pivoted[column] = ''
    pivoted = pivoted[LEGACY_ACTIVITY_EXPORT_COLUMNS]
    flat_rows = pivoted.to_dict('records')

    preview_rows = []
    for row in flat_rows[:6]:
        preview_rows.append({
            'date': row.get('Activity Date', ''),
            'store': row.get('Retailer Name', ''),
            'city': row.get('Retailer City', ''),
            'inventory': row.get('Store Inventory?', '') or 'Not stated',
            'opportunity': row.get('Opportunities or Concerns?', '') or 'No note captured',
        })
    return {
        'rows': ordered,
        'count': len(ordered),
        'pdf_rows': ordered[:20],
        'flat_rows': flat_rows,
        'preview_rows': preview_rows,
    }


def _build_executive_summary(brand_name, current, issue_mix, opportunity_mix, geography, sales_row, previous_metrics, trailing_avg, quality_label):
    summary_lines = [
        f"{brand_name} logged {current['activities_logged']:,} field activities across {current['stores_visited']} retailers in {current['states_covered']} states and {current['cities_covered']} cities during the period."
    ]
    if sales_row:
        revenue = _safe_float(sales_row.get('total_revenue'))
        sales_stores = _safe_int(sales_row.get('num_stores'))
        repeat_pct = _safe_float(sales_row.get('repeat_pct'))
        summary_lines.append(f"Commercially, the brand closed ₦{revenue:,.2f} from {sales_stores} selling stores with repeat purchase at {repeat_pct:.1f}%.")
    if issue_mix:
        summary_lines.append(f"The most frequent issue was {issue_mix[0]['label'].lower()} ({issue_mix[0]['count']} mentions).")
    if opportunity_mix:
        summary_lines.append(f"The strongest opportunity signal was {opportunity_mix[0]['label']} ({opportunity_mix[0]['count']} mentions).")
    if quality_label != 'Full fidelity':
        summary_lines.append('Salesperson-level detail is partially unavailable because this batch came from the cleaned weekly pack.')

    bullets = [
        f"Coverage concentrated most in {geography['top_state']} / {geography['top_city']}.",
        f"{current['issues_found']:,} issues and {current['opportunities_found']:,} opportunity signals were logged.",
    ]
    if previous_metrics:
        prev_current = previous_metrics['current']
        change = _calc_pct_change(current['activities_logged'], prev_current.get('activities_logged'))
        if change is not None:
            bullets.append(f"Activity volume moved {_format_delta(change)} versus the previous comparable week.")
    if trailing_avg.get('activities_logged') is not None:
        change = _calc_pct_change(current['activities_logged'], trailing_avg['activities_logged'])
        if change is not None:
            bullets.append(f"Current activity sits {_format_delta(change)} against the trailing 4-week average.")
    if sales_row and _safe_float(sales_row.get('repeat_pct')) < 35:
        bullets.append('Low repeat purchase means field momentum is not yet translating strongly into reorder behavior.')
    return {'paragraph': ' '.join(summary_lines), 'bullets': bullets[:6]}


def _build_recommendations(current, issue_mix, opportunity_mix, store_breakdown, quality_label, sales_row):
    what_stands_out, risks, opportunities = [], [], []
    do_now, watch, strategic = [], [], []

    if issue_mix:
        what_stands_out.append(f"{issue_mix[0]['label']} is the lead issue signal with {issue_mix[0]['count']} mentions.")
        risks.append(f"{issue_mix[0]['label']} is still recurring across the field network.")
        do_now.append({
            'action': f"Resolve {issue_mix[0]['label'].lower()} hotspots immediately.",
            'why': f"It was the most frequent issue recorded this period ({issue_mix[0]['count']} mentions).",
            'affected': ', '.join(row['retailer_name'] for row in store_breakdown['rows'][:3]),
            'impact': 'Reduces preventable friction before the next sales cycle.',
            'urgency': 'High',
        })
    if opportunity_mix:
        what_stands_out.append(f"{opportunity_mix[0]['label']} is the strongest opportunity signal right now.")
        opportunities.append(f"Field teams repeatedly flagged {opportunity_mix[0]['label']} as an upside pocket.")
        strategic.append({
            'action': f"Convert the strongest {opportunity_mix[0]['label']} opportunities into commercial asks.",
            'why': 'The same opportunity pattern is repeating across multiple store visits.',
            'affected': ', '.join(row['retailer_name'] for row in store_breakdown['rows'][:5]),
            'impact': 'Improves conversion from activity into distribution or reorder growth.',
            'urgency': 'Medium',
        })

    if store_breakdown['rows']:
        top_issue_store = sorted(store_breakdown['rows'], key=lambda row: row.get('issues', 0), reverse=True)[0]
        top_opportunity_store = sorted(store_breakdown['rows'], key=lambda row: row.get('opportunities', 0), reverse=True)[0]
        risks.append(f"{top_issue_store['retailer_name']} logged the highest issue load ({top_issue_store['issues']} issues).")
        opportunities.append(f"{top_opportunity_store['retailer_name']} carried the strongest opportunity load ({top_opportunity_store['opportunities']} signals).")
        watch.append({
            'action': f"Track issue closure in {top_issue_store['retailer_name']}.",
            'why': 'It is the highest-risk store in the latest activity footprint.',
            'affected': top_issue_store['retailer_name'],
            'impact': 'Prevents one problematic outlet from distorting brand perception.',
            'urgency': 'Medium',
        })

    if sales_row:
        repeat_pct = _safe_float(sales_row.get('repeat_pct'))
        if repeat_pct < 35:
            do_now.append({
                'action': 'Link field follow-up to repeat-purchase recovery.',
                'why': f'Repeat purchase is only {repeat_pct:.1f}% despite field coverage.',
                'affected': 'Top selling stores with weak repeat behavior',
                'impact': 'Improves conversion from visibility to reorder.',
                'urgency': 'High',
            })
        elif repeat_pct >= 60:
            opportunities.append('Repeat purchase is healthy enough to support deeper assortment conversations.')

    if quality_label != 'Full fidelity':
        watch.append({
            'action': 'Pair the cleaned weekly pack with the raw export when rep accountability matters.',
            'why': 'The cleaned pack does not preserve full salesperson metadata.',
            'affected': 'Field team analysis',
            'impact': 'Prevents under-reporting of rep-level contribution.',
            'urgency': 'Low',
        })

    if not do_now:
        do_now.append({
            'action': 'Protect the strongest stores and close the biggest issue pockets first.',
            'why': 'Current data shows active field momentum but still exposed exceptions.',
            'affected': 'Top activity stores',
            'impact': 'Keeps activity effort tied to near-term commercial outcomes.',
            'urgency': 'Medium',
        })
    return {
        'what_stands_out': what_stands_out[:4],
        'risks': risks[:4],
        'opportunities': opportunities[:4],
        'actions': {'do_now': do_now[:3], 'watch': watch[:3], 'strategic': strategic[:3]},
    }


def prepare_activity_report_data(ds, brand_name: str, report_id: int = None, batch_id: int = None,
                                 start_date: str = None, end_date: str = None,
                                 report_type: str = None, period_label: str = None) -> dict:
    report = ds.get_report(report_id) if report_id else None
    batch_row, batch_summary = _get_activity_batch_context(ds, report_id=report_id, batch_id=batch_id)
    if not report and not batch_row:
        fallback_report = ds.get_latest_report() if not batch_id else None
        report = fallback_report
        if fallback_report:
            batch_row, batch_summary = _get_activity_batch_context(ds, report_id=fallback_report.get('id'))
    if not report and not batch_row:
        raise ValueError('No report or activity batch found for activity report generation')

    context_type = report.get('report_type') if report else batch_summary.get('report_type') or report_type or 'custom'
    context_label = (report.get('month_label') if report else None) or batch_summary.get('report_label') or period_label or 'Current Period'
    context_start = start_date or (report.get('start_date') if report else None) or batch_summary.get('start_date')
    context_end = end_date or (report.get('end_date') if report else None) or batch_summary.get('end_date')
    source_type_label, quality_label, quality_flags = _source_labels(batch_row, batch_summary)

    events_df, visits_df, issues_df, mentions_df = _read_activity_frames(ds, brand_name, report_id=report_id, batch_id=batch_id)
    current, non_opportunity_issues, opportunity_issues = _compute_current_metrics(events_df, visits_df, issues_df, mentions_df)
    previous = previous_metrics = same_period_last_year = yoy_metrics = None
    trailing = []
    trailing_metrics = []
    if report:
        previous, previous_metrics, trailing, trailing_metrics, same_period_last_year, yoy_metrics = _build_comparisons(ds, brand_name, report)
    previous_events_df = pd.DataFrame()
    previous_visits_df = pd.DataFrame()
    if previous:
        previous_events_df, previous_visits_df, _, _ = _read_activity_frames(ds, brand_name, report_id=previous['id'])
    period_breakdown = build_period_comparison(
        _build_period_rows(visits_df),
        _build_period_rows(previous_visits_df),
        date_key='date',
        metric_columns=('activities', 'issues', 'opportunities'),
        report_type=context_type,
        current_start=context_start,
        previous_start=previous.get('start_date') if previous else None,
    ) if not events_df.empty else None
    kpi_cards, trailing_avg = _build_kpi_cards(current, previous_metrics, trailing_metrics, quality_label)
    issue_mix, opportunity_mix, highlights = _build_mix_sections(non_opportunity_issues, opportunity_issues, visits_df)
    store_breakdown = _build_store_breakdown(visits_df, mentions_df)
    geography = _build_geography(events_df)
    field_team = _build_field_team(events_df, visits_df, quality_label)
    details = _build_details(events_df)
    sales_row = ds.get_brand_kpis_single(report_id, brand_name) if report_id else None
    executive_summary = _build_executive_summary(brand_name, current, issue_mix, opportunity_mix, geography, sales_row, previous_metrics, trailing_avg, quality_label)
    recommendations = _build_recommendations(current, issue_mix, opportunity_mix, store_breakdown, quality_label, sales_row)

    previous_current = previous_metrics['current'] if previous_metrics else {}
    yoy_current = yoy_metrics['current'] if yoy_metrics else {}
    comparison = {
        'grain': comparison_basis_label(context_type),
        'basis': [
            f"Vs previous comparable {context_type or 'period'}" + (f" ({previous.get('month_label')})" if previous else ' (not available)'),
            "Vs trailing 4-week average" if trailing_metrics else "Vs trailing 4-week average (not available)",
            f"Vs same period last year ({same_period_last_year.get('month_label')})" if same_period_last_year else "Vs same period last year (not available)",
        ],
        'previous': {
            'label': previous.get('month_label') if previous else 'Not available',
            'activities_change': _calc_pct_change(current['activities_logged'], previous_current.get('activities_logged')),
            'issues_change': _calc_pct_change(current['issues_found'], previous_current.get('issues_found')),
            'opportunities_change': _calc_pct_change(current['opportunities_found'], previous_current.get('opportunities_found')),
            'stores_change': _calc_pct_change(current['stores_visited'], previous_current.get('stores_visited')),
        },
        'trailing': {
            'activities_change': _calc_pct_change(current['activities_logged'], trailing_avg.get('activities_logged')),
            'issues_change': _calc_pct_change(current['issues_found'], trailing_avg.get('issues_found')),
            'opportunities_change': _calc_pct_change(current['opportunities_found'], trailing_avg.get('opportunities_found')),
            'stores_change': _calc_pct_change(current['stores_visited'], trailing_avg.get('stores_visited')),
        },
        'yoy': {
            'label': same_period_last_year.get('month_label') if same_period_last_year else 'Not available',
            'activities_change': _calc_pct_change(current['activities_logged'], yoy_current.get('activities_logged')),
            'issues_change': _calc_pct_change(current['issues_found'], yoy_current.get('issues_found')),
            'opportunities_change': _calc_pct_change(current['opportunities_found'], yoy_current.get('opportunities_found')),
            'stores_change': _calc_pct_change(current['stores_visited'], yoy_current.get('stores_visited')),
        }
    }
    comparison['period_breakdown'] = period_breakdown
    comparison['cards'] = [
        _build_comparison_card(
            'Previous Comparable',
            comparison['basis'][0],
            comparison['previous'],
            comparison['previous']['label'],
        ),
        _build_comparison_card(
            'Trailing 4-Week Average',
            comparison['basis'][1],
            comparison['trailing'],
            'Last 4 comparable periods' if trailing_metrics else 'Not available',
        ),
        _build_comparison_card(
            'Same Period Last Year',
            comparison['basis'][2],
            comparison['yoy'],
            comparison['yoy']['label'],
        ),
    ]

    return {
        'identity': {
            'brand_name': brand_name,
            'report_title': 'Field Activity Report',
            'period_label': context_label,
            'date_range': _compact_date_label(context_start, context_end),
            'generated_at': datetime.now().strftime('%d %b %Y %I:%M %p'),
            'source_type': source_type_label,
            'source_quality': quality_label,
            'quality_flags': quality_flags,
            'report_id': report_id,
            'batch_id': batch_id,
            'report_type': context_type,
            'start_date': context_start,
            'end_date': context_end,
            'source_filename': (batch_row or {}).get('source_filename'),
        },
        'executive_summary': executive_summary,
        'kpis': current,
        'kpi_cards': kpi_cards,
        'comparison': comparison,
        'sales_context': {
            'available': bool(sales_row),
            'revenue': _safe_float(sales_row.get('total_revenue')) if sales_row else 0,
            'quantity': _safe_float(sales_row.get('total_qty')) if sales_row else 0,
            'sales_stores': _safe_int(sales_row.get('num_stores')) if sales_row else 0,
            'repeat_pct': _safe_float(sales_row.get('repeat_pct')) if sales_row else 0,
        },
        'activity_trend': _build_daily_trend(visits_df),
        'issues_opportunities': {
            'issue_mix': issue_mix,
            'opportunity_mix': opportunity_mix,
            'highlights': highlights,
        },
        'store_breakdown': store_breakdown,
        'geography': geography,
        'field_team': field_team,
        'details': details,
        'recommended': recommendations,
        'source_manifest': {
            'row_count': batch_summary.get('row_count'),
            'file_count': batch_summary.get('file_count') or (batch_summary.get('source_meta') or {}).get('file_count'),
            'start_date': batch_summary.get('start_date'),
            'end_date': batch_summary.get('end_date'),
            'source_meta': batch_summary.get('source_meta') or {},
        }
    }


def _render_activity_report_html(activity_data: dict, *, include_full_details: bool, for_pdf: bool) -> str:
    template = jinja_env.get_template('activity_report_dashboard.html')
    context = dict(activity_data)
    context['chart_json'] = json.dumps({
        'trend': activity_data['activity_trend'],
        'issue_mix': activity_data['issues_opportunities']['issue_mix'],
        'opportunity_mix': activity_data['issues_opportunities']['opportunity_mix'],
        'stores_activity': activity_data['store_breakdown']['activity_chart'],
        'stores_issues': activity_data['store_breakdown']['issues_chart'],
        'stores_opportunities': activity_data['store_breakdown']['opportunity_chart'],
        'states': activity_data['geography']['states'],
        'cities': activity_data['geography']['cities'],
        'field_activity': activity_data['field_team']['activity_chart'],
        'field_issues': activity_data['field_team']['issue_chart'],
        'field_stores': activity_data['field_team']['store_chart'],
    })
    context['details_rows'] = activity_data['details']['rows'] if include_full_details else activity_data['details']['pdf_rows']
    context['include_full_details'] = include_full_details
    context['for_pdf'] = for_pdf
    return template.render(**context)


def generate_activity_report_html(output_path: str, brand_name: str, activity_data: dict, period_label: str = None,
                                  report_id: int = None, html_output_path: str = None) -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if not html_output_path:
        html_output_path = output_path.replace(f"{os.sep}pdf{os.sep}", f"{os.sep}html{os.sep}").replace('.pdf', '.html')
    os.makedirs(os.path.dirname(html_output_path), exist_ok=True)
    html_content = _render_activity_report_html(activity_data, include_full_details=True, for_pdf=False)
    with open(html_output_path, 'w', encoding='utf-8') as handle:
        handle.write(html_content)

    pdf_content = _render_activity_report_html(activity_data, include_full_details=False, for_pdf=True)
    try:
        pdf_bytes = render_pdf_bytes(pdf_content)
        with open(output_path, 'wb') as handle:
            handle.write(pdf_bytes)
        return output_path
    except Exception:
        return html_output_path


def generate_activity_excel_report(output_path: str, activity_data: dict) -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    headers = LEGACY_ACTIVITY_EXPORT_COLUMNS
    rows = activity_data['details'].get('flat_rows') or []
    ws.append(headers)
    if rows:
        for row in rows:
            ws.append([row.get(header, '') for header in headers])
    else:
        ws.append(['No activity detail available'] + [''] * (len(headers) - 1))

    fill = PatternFill('solid', fgColor='1B2B5E')
    for cell in ws[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    widths = {
        'A': 16, 'B': 34, 'C': 18, 'D': 18, 'E': 28, 'F': 24, 'G': 30,
        'H': 42, 'I': 16, 'J': 32, 'K': 28, 'L': 18, 'M': 18, 'N': 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    wb.save(output_path)
    return output_path
