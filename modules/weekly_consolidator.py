from __future__ import annotations

import io
import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import Workbook, load_workbook


MASTER_COLUMNS = [
    "Brand Partners",
    "Particulars",
    "Date",
    "Retailers",
    "Vch Type",
    "Vch No.",
    "Quantity",
    "Sales_Value",
]

SHEET_NAME = "Itemwise Stock Details"

BRAND_ALIASES = {
    "August Secrets Limted": "August Secrets",
    "August Secrets Limited": "August Secrets",
    "Whole Eats Africa": "WholeEats",
    "Madala Beverages Limited": "Madala Beverages",
    "Wilson's Juice Co.Ltd": "Wilson's",
    "Zayith Food Company": "Zayith",
    "Mazara Foods": "Mazara",
    "Cressolife Service Ltd": "Cressolife JT",
    "Orisirisi Catring Service": "Orisirisi",
}

PREFIX_BRAND = {
    "AGS": "August Secrets",
    "AMB": "Aman Blessed",
    "AMY": "Amayi Foods",
    "BQ": "Bboon Foods",
    "BNN": "Biniowan Enterprises",
    "MAD": "Madala Beverages",
    "WF": "WholeEats",
    "ZY": "Zayith",
    "ZF": "Zeef Food",
    "KS": "Kezia Foods",
    "TF": "Tosh Cocodia",
    "CN": "Cressolife JT",
}


@dataclass
class WeeklyConsolidationResult:
    dataframe: pd.DataFrame
    start_date: date
    end_date: date
    output_filename: str
    source_counts: dict[str, int]


def _clean_text(value) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).replace("_x000D_\n", " ").replace("_x000D_", " ")
    text = text.replace("\r", " ").replace("\n", " ")
    return " ".join(text.split()).strip()


def _strip_bp(value) -> str:
    text = re.sub(r"^BP\s+", "", _clean_text(value), flags=re.I).strip()
    return BRAND_ALIASES.get(text, text)


def _brand_from_sku(sku: str, fallback: str = "") -> str:
    match = re.match(r"^([A-Z]{2,5})-", _clean_text(sku))
    if match and match.group(1) in PREFIX_BRAND:
        return PREFIX_BRAND[match.group(1)]
    return _strip_bp(fallback) if fallback else ""


def _is_sku_label(value) -> bool:
    text = _clean_text(value)
    if not text:
        return False
    if re.match(r"^[A-Z]{2,5}-", text):
        return True
    if re.search(r"\(\d+x\)", text, flags=re.I):
        return True
    return bool(re.search(r"\b\d+(?:\.\d+)?\s*(?:g|kg|ml|l|ltr|litre)\b", text, flags=re.I))


def _read_excel(file_source, sheet_name: str, header=None) -> pd.DataFrame:
    if hasattr(file_source, "seek"):
        file_source.seek(0)
    return pd.read_excel(file_source, sheet_name=sheet_name, header=header)


def _coerce_master(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy() if df is not None else pd.DataFrame()
    for column in MASTER_COLUMNS:
        if column not in out.columns:
            out[column] = pd.NA
    out = out[MASTER_COLUMNS].copy()
    for column in ["Brand Partners", "Particulars", "Retailers", "Vch Type", "Vch No."]:
        out[column] = out[column].apply(_clean_text)
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce")
    out["Quantity"] = pd.to_numeric(out["Quantity"], errors="coerce")
    out["Sales_Value"] = pd.to_numeric(out["Sales_Value"], errors="coerce")
    return out


def _parse_itemwise(file_source) -> pd.DataFrame:
    raw = _read_excel(file_source, SHEET_NAME, header=None)
    header_idx = 1
    for idx, row in raw.iterrows():
        values = [_clean_text(value).lower() for value in row.tolist()]
        if "vch type" in values and "sales_value" in values:
            header_idx = idx
            break

    data = raw.iloc[header_idx + 1 :].copy().reset_index(drop=True)
    if 0 in data.columns:
        data = data.drop(columns=[0])
    data = data.rename(
        columns={
            1: "Brand Partners",
            2: "Particulars",
            3: "Date",
            4: "Retailers",
            5: "Vch Type",
            6: "Vch No.",
            7: "Quantity",
            8: "Sales_Value",
        }
    )
    data = _coerce_master(data).dropna(how="all")
    # Journal rows are rebuilt from the journal register because that file carries
    # the useful destination narration; keeping both would double count.
    keep = ["Sales", "Inventory Pickup by Dala", "Inventory Supplied by Brands"]
    return data[data["Vch Type"].isin(keep)].reset_index(drop=True)


def _parse_inventory(file_source) -> pd.DataFrame:
    raw = _read_excel(file_source, "Stock Category Summary", header=None)
    report_date = pd.to_datetime(_clean_text(raw.iloc[1, 0]).replace("For ", ""), format="%d-%b-%y", errors="coerce")
    vch_no = f"Inv:{report_date.day}{report_date.month}{str(report_date.year)[-2:]}" if pd.notna(report_date) else "Inv:"
    current_brand = ""
    records: list[dict] = []

    for _, row in raw.iloc[6:].iterrows():
        first = _clean_text(row.iloc[0] if len(row) else "")
        if not first:
            continue
        lowered = first.lower()
        if any(marker in lowered for marker in ("grand total", "printed by", "order generation", "head office")):
            break
        quantity = pd.to_numeric(row.iloc[2] if len(row) > 2 else None, errors="coerce")
        value = pd.to_numeric(row.iloc[3] if len(row) > 3 else None, errors="coerce")
        if not _is_sku_label(first):
            current_brand = _strip_bp(first)
            continue
        if pd.isna(quantity) and pd.isna(value):
            continue
        brand = current_brand or _brand_from_sku(first)
        records.append(
            {
                "Brand Partners": brand,
                "Particulars": first,
                "Date": report_date,
                "Retailers": brand,
                "Vch Type": "Available Inventory",
                "Vch No.": vch_no,
                "Quantity": quantity,
                "Sales_Value": value,
            }
        )
    return _coerce_master(pd.DataFrame(records, columns=MASTER_COLUMNS))


def _split_voucher_groups(raw: pd.DataFrame) -> list[list[pd.Series]]:
    groups: list[list[pd.Series]] = []
    current: list[pd.Series] = []
    for _, row in raw.iloc[3:].iterrows():
        if pd.notna(row.iloc[0]):
            if current:
                groups.append(current)
            current = [row]
        elif current:
            current.append(row)
    if current:
        groups.append(current)
    return groups


def _extract_retailer_from_narration(text: str) -> str:
    cleaned = _clean_text(text)
    patterns = [
        r"(?i)\b(?:supplied|delivered|dispatch(?:ed)?|sent)\s+(?:to|for)\s+(.+?)(?:\s+on\b|\s+by\b|,|$)",
        r"(?i)\bgoods\s+(?:supplied|delivered)\s+to\s+(.+?)(?:\s+on\b|,|$)",
        r"(?i)\bto\s+(.+?)(?:\s+on\b|,|$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, cleaned)
        if match:
            return _clean_text(match.group(1)).upper()
    return cleaned.upper() if cleaned else "-- CHECK RETAILER --"


def _parse_journal(file_source) -> pd.DataFrame:
    raw = _read_excel(file_source, "Journal Register", header=None)
    records: list[dict] = []
    for group in _split_voucher_groups(raw):
        header = group[0]
        voucher_date = pd.to_datetime(header.iloc[0], errors="coerce")
        header_brand = _strip_bp(header.iloc[1] if len(header) > 1 else "")
        vch_no = _clean_text(header.iloc[5] if len(header) > 5 else "")
        narration_parts: list[str] = []
        items: list[tuple[str, float, float]] = []
        for row in group[1:]:
            text = _clean_text(row.iloc[1] if len(row) > 1 else "")
            if not text or text.lower() in {"out-going stock", "out going stock", "stock"}:
                continue
            quantity = pd.to_numeric(row.iloc[2] if len(row) > 2 else None, errors="coerce")
            amount = pd.to_numeric(row.iloc[4] if len(row) > 4 else None, errors="coerce")
            if _is_sku_label(text) and pd.notna(quantity):
                items.append((text, quantity, amount))
            else:
                narration_parts.append(text)
        retailer = _extract_retailer_from_narration(" ".join(narration_parts))
        for sku, quantity, amount in items:
            records.append(
                {
                    "Brand Partners": _brand_from_sku(sku, header_brand) or header_brand,
                    "Particulars": sku,
                    "Date": voucher_date,
                    "Retailers": retailer,
                    "Vch Type": "Journal",
                    "Vch No.": vch_no,
                    "Quantity": quantity,
                    "Sales_Value": amount,
                }
            )
    return _coerce_master(pd.DataFrame(records, columns=MASTER_COLUMNS))


def _parse_credit(file_source) -> pd.DataFrame:
    raw = _read_excel(file_source, "Credit Note Register", header=None)
    records: list[dict] = []
    for group in _split_voucher_groups(raw):
        header = group[0]
        brand = ""
        for row in group[1:]:
            text = _clean_text(row.iloc[1] if len(row) > 1 else "")
            if text.lower().startswith("bp "):
                brand = _strip_bp(text)
                break
        brand = brand or "Unknown"
        records.append(
            {
                "Brand Partners": brand,
                "Particulars": brand,
                "Date": pd.to_datetime(header.iloc[0], errors="coerce"),
                "Retailers": _clean_text(header.iloc[1] if len(header) > 1 else ""),
                "Vch Type": "Credit Note",
                "Vch No.": _clean_text(header.iloc[5] if len(header) > 5 else ""),
                "Quantity": 0.0,
                "Sales_Value": pd.to_numeric(header.iloc[7] if len(header) > 7 else None, errors="coerce"),
            }
        )
    return _coerce_master(pd.DataFrame(records, columns=MASTER_COLUMNS))


def _excel_safe_source(file_source):
    if isinstance(file_source, (str, Path)):
        return file_source
    if hasattr(file_source, "read"):
        data = file_source.read()
        return io.BytesIO(data)
    return file_source


def consolidate_weekly_files(*, sales_file, inventory_file, journal_file, credit_file, output_filename: str | None = None) -> WeeklyConsolidationResult:
    sources = {
        "itemwise": _parse_itemwise(_excel_safe_source(sales_file)),
        "journal": _parse_journal(_excel_safe_source(journal_file)),
        "credit": _parse_credit(_excel_safe_source(credit_file)),
        "inventory": _parse_inventory(_excel_safe_source(inventory_file)),
    }
    master = pd.concat(sources.values(), ignore_index=True)
    master = _coerce_master(master)
    master = master.sort_values(["Brand Partners", "Date", "Particulars", "Vch Type"], kind="stable").reset_index(drop=True)
    dates = master["Date"].dropna()
    start_date = dates.min().date()
    end_date = dates.max().date()
    filename = output_filename or f"Weekly Sales Report {start_date.isoformat()} to {end_date.isoformat()}.xlsx"
    return WeeklyConsolidationResult(
        dataframe=master,
        start_date=start_date,
        end_date=end_date,
        output_filename=filename,
        source_counts={name: len(frame) for name, frame in sources.items()},
    )


def write_consolidated_workbook(df: pd.DataFrame, template_path: str | Path | None = None) -> bytes:
    template = Path(template_path) if template_path else None
    if template and template.exists():
        template_wb = load_workbook(template)
        template_ws = template_wb[SHEET_NAME]
    else:
        template_ws = None

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    default_widths = [25.6640625, 44.5546875, 9.21875, 42.33203125, 25.21875, 9.88671875, 9.5546875, 13.109375]
    for col_idx, column in enumerate(MASTER_COLUMNS, start=1):
        letter = ws.cell(1, col_idx).column_letter
        ws.column_dimensions[letter].width = (
            template_ws.column_dimensions[letter].width if template_ws is not None else default_widths[col_idx - 1]
        )
        cell = ws.cell(1, col_idx, column)
        if template_ws is not None:
            _copy_cell_style(template_ws.cell(1, col_idx), cell)
        else:
            font = copy(cell.font)
            font.bold = True
            cell.font = font

    frame = _coerce_master(df)
    for row_idx, row in enumerate(frame.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            if col_idx == 3 and pd.notna(value):
                value = pd.Timestamp(value).to_pydatetime()
            elif pd.isna(value):
                value = None
            cell = ws.cell(row_idx, col_idx, value)
            if template_ws is not None:
                _copy_cell_style(template_ws.cell(2, col_idx), cell)
            elif col_idx == 3:
                cell.number_format = "d-mmm-yy"
            elif col_idx in {7, 8}:
                cell.number_format = '#,##0.00'

    ws.auto_filter.ref = f"A1:H{len(frame) + 1}"
    ws.freeze_panes = None
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _copy_cell_style(source, target) -> None:
    if not source.has_style:
        return
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)
